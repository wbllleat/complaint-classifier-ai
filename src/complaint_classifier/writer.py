"""Write classification results back to Excel."""

from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from .models import ClassifyResult, ComplaintRow

LOW_CONFIDENCE_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def write_results(
    input_path: str,
    output_path: str,
    rows: list[ComplaintRow],
    results: list[ClassifyResult],
) -> str:
    """Write classification results to a new Excel file.

    Only writes to rows that have corresponding classification results.
    Rows without results are left unchanged.

    Args:
        input_path: Path to original Excel file.
        output_path: Path for output Excel file.
        rows: All complaint rows from the input.
        results: Classification results (subset is OK).

    Returns:
        Output file path.
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("No active worksheet")

    # Add new header columns (only once)
    last_col = ws.max_column
    new_headers = ["一级分类", "二级分类", "置信度", "分类理由"]
    for i, header in enumerate(new_headers, start=1):
        ws.cell(row=1, column=last_col + i, value=header)

    # Build lookup from row_index to result
    result_map: dict[int, ClassifyResult] = {r["row_index"]: r for r in results}

    # Write results only for rows that have a match
    for row in ws.iter_rows(min_row=2):
        excel_row_num = row[0].row
        result = result_map.get(excel_row_num)
        if result is None:
            continue
        ws.cell(row=excel_row_num, column=last_col + 1, value=result["primary_category"])
        ws.cell(row=excel_row_num, column=last_col + 2, value=result["secondary_category"])
        ws.cell(row=excel_row_num, column=last_col + 3, value=result["confidence"])
        ws.cell(row=excel_row_num, column=last_col + 4, value=result["reasoning"])

        if result["confidence"] == "low":
            for col_offset in range(1, 5):
                cell = ws.cell(row=excel_row_num, column=last_col + col_offset)
                cell.fill = copy(LOW_CONFIDENCE_FILL)

    wb.save(output_path)
    wb.close()
    return str(output_path)


def write_test_results(
    output_path: str,
    rows: list[ComplaintRow],
    results: list[ClassifyResult],
) -> str:
    """Write test sample results to a standalone Excel file.

    Creates a new workbook with only the classified sample rows.

    Args:
        output_path: Path for output Excel file.
        rows: Classified sample rows.
        results: Classification results (same length as rows).

    Returns:
        Output file path.
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
    ws.title = "测试分类结果"

    # Headers
    headers = [
        "行号", "工单号", "投诉主题", "受理渠道",
        "归档意见(前200字)", "一级分类", "二级分类", "置信度", "分类理由",
    ]
    ws.append(headers)

    result_map: dict[int, ClassifyResult] = {r["row_index"]: r for r in results}

    for row in rows:
        result = result_map.get(row.row_index)
        if result is None:
            continue
        ws.append([
            row.row_index,
            row.工单号,
            row.投诉主题,
            row.受理渠道,
            row.归档意见[:200],
            result["primary_category"],
            result["secondary_category"],
            result["confidence"],
            result["reasoning"],
        ])
        # Highlight low confidence rows
        if result["confidence"] == "low":
            row_num = ws.max_row
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).fill = copy(LOW_CONFIDENCE_FILL)

    wb.save(output_path)
    wb.close()
    return str(output_path)
