"""Excel file reading and column detection."""

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import ColumnNotFoundError, ComplaintRow


def read_excel(path: str, sheet_name: str | None) -> tuple[list[str], list[ComplaintRow]]:
    """Read Excel file and return headers and parsed rows.

    Args:
        path: Path to .xlsx file.
        sheet_name: Sheet name to read, or None for active sheet.

    Returns:
        Tuple of (headers list, complaint rows list).

    Raises:
        FileNotFoundError: If file does not exist.
        ColumnNotFoundError: If archive column not detected.
    """
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = load_workbook(str(file_path), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    if ws is None:
        raise ValueError(f"No active worksheet in {path}")

    # Read headers from row 1
    headers = [_safe_str(cell.value) for cell in ws[1]]

    # Read data rows (row 2 onwards)
    rows: list[ComplaintRow] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty_row(row):
            continue
        complaint = _parse_row(row_idx, headers, row)
        rows.append(complaint)

    wb.close()
    return headers, rows


def detect_archive_column(headers: list[str], target_name: str) -> int:
    """Find the index of the archive opinion column.

    Args:
        headers: List of column header strings.
        target_name: Expected column name.

    Returns:
        Zero-based column index.

    Raises:
        ColumnNotFoundError: If no match found.
    """
    # Exact match first
    for i, h in enumerate(headers):
        if h == target_name:
            return i
    # Fuzzy match: target_name is substring of header
    for i, h in enumerate(headers):
        if target_name in h:
            return i
    raise ColumnNotFoundError(
        f"Column '{target_name}' not found in headers: {headers}"
    )


def get_sheet_names(path: str) -> list[str]:
    """Get all worksheet names from an Excel file.

    Args:
        path: Path to .xlsx file.

    Returns:
        List of sheet names.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _safe_str(value: Any) -> str:
    """Convert cell value to string, handling None."""
    if value is None:
        return ""
    return str(value)


def _parse_row(row_idx: int, headers: list[str], row: tuple) -> ComplaintRow:
    """Parse a single row tuple into a ComplaintRow."""
    values = {h: v for h, v in zip(headers, row) if h}
    return ComplaintRow(
        row_index=row_idx,
        工单号=_safe_str(values.get("工单号", "")),
        投诉主题=_safe_str(values.get("投诉主题", "")),
        号码=_safe_str(values.get("号码", "")),
        受理时间=_parse_datetime(values.get("受理时间")),
        投诉内容=_safe_str(values.get("投诉内容", "")),
        工单状态=_safe_str(values.get("工单状态", "")),
        受理渠道=_safe_str(values.get("受理渠道", "")),
        归档意见=_safe_str(values.get("归档意见", "")),
        归档时间=_parse_datetime(values.get("归档时间")),
        联系电话=_safe_str(values.get("联系电话", "")),
        客户星级=_safe_str(values.get("客户星级", "")),
        受理部门=_safe_str(values.get("受理部门", "")),
        受理员工=_safe_str(values.get("受理员工", "")),
        客户归属地=_safe_str(values.get("客户归属地", "")),
        业务地市=_safe_str(values.get("业务地市", "")),
    )


def _parse_datetime(value: Any) -> datetime | None:
    """Parse cell value to datetime or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    return None


def _is_empty_row(row: tuple) -> bool:
    """Check if a row is entirely empty."""
    return all(v is None for v in row)
